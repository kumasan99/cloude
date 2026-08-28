"""data/inbox に置かれたファイルを読み取り、テキストとして取り出す。

対応形式: .txt .md .csv .tsv .json .yaml .yml .xlsx .pdf
それ以外は「読めなかったファイル」として一覧に残し、CEOに変換をお願いする。
"""

import csv
import fnmatch
import io
import json
from datetime import datetime, timezone
from pathlib import Path

from .config import CATEGORIES, INBOX_DIR, data_requirements

# 1ファイルあたりの読み取り上限（文字数）。巨大な台帳で文脈を食い潰さないため。
MAX_CHARS_PER_FILE = 40_000
# 全ファイル合計の上限
MAX_TOTAL_CHARS = 300_000

TEXT_SUFFIXES = {".txt", ".md", ".json", ".yaml", ".yml", ".log"}
TABLE_SUFFIXES = {".csv", ".tsv"}


def _read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def _read_table(path: Path) -> str:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    rows = list(csv.reader(io.StringIO(raw), delimiter=delimiter))
    lines = [" | ".join(cell.strip() for cell in row) for row in rows if any(row)]
    return "\n".join(lines)


def _read_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    chunks = []
    for sheet in wb.worksheets:
        chunks.append(f"## シート: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            chunks.append(" | ".join("" if c is None else str(c) for c in row))
    wb.close()
    return "\n".join(chunks)


def _read_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    return "\n".join(
        f"--- p.{i + 1} ---\n{(page.extract_text() or '').strip()}"
        for i, page in enumerate(reader.pages)
    )


def _extract(path: Path) -> tuple[str, str]:
    """(text, status) を返す。status は ok / unsupported / error。"""
    suffix = path.suffix.lower()
    try:
        if suffix in TEXT_SUFFIXES:
            return _read_text(path), "ok"
        if suffix in TABLE_SUFFIXES:
            return _read_table(path), "ok"
        if suffix == ".xlsx":
            return _read_xlsx(path), "ok"
        if suffix == ".pdf":
            return _read_pdf(path), "ok"
    except Exception as exc:  # 読めなくても他のファイルの処理は続ける
        return f"（読み取りエラー: {exc}）", "error"
    return "", "unsupported"


def scan_inbox() -> list[dict]:
    """inbox 配下のファイル一覧をメタ情報つきで返す（本文は含まない）。"""
    files = []
    if not INBOX_DIR.exists():
        return files
    for path in sorted(INBOX_DIR.rglob("*")):
        if not path.is_file() or path.name.startswith("."):
            continue
        rel = path.relative_to(INBOX_DIR)
        category = rel.parts[0] if len(rel.parts) > 1 else "uncategorized"
        stat = path.stat()
        files.append(
            {
                "path": str(rel).replace("\\", "/"),
                "name": path.name,
                "category": category,
                "category_label": CATEGORIES.get(category, "未分類"),
                "size_bytes": stat.st_size,
                "modified_at": datetime.fromtimestamp(
                    stat.st_mtime, tz=timezone.utc
                ).astimezone().isoformat(timespec="seconds"),
                "suffix": path.suffix.lower(),
            }
        )
    return files


def load_documents() -> list[dict]:
    """inbox のファイルを読み取り、本文つきで返す。"""
    documents = []
    total = 0
    for meta in scan_inbox():
        path = INBOX_DIR / meta["path"]
        text, status = _extract(path)
        truncated = False
        if len(text) > MAX_CHARS_PER_FILE:
            text = text[:MAX_CHARS_PER_FILE]
            truncated = True
        if total + len(text) > MAX_TOTAL_CHARS:
            text = text[: max(0, MAX_TOTAL_CHARS - total)]
            truncated = True
        total += len(text)
        documents.append({**meta, "status": status, "truncated": truncated, "text": text})
    return documents


def requirement_status(files: list[dict] | None = None) -> list[dict]:
    """必要資料チェックリストに対して、提供済み / 未提供を判定する。

    判定は2経路。
    - ローカル: data/inbox のファイル名が match パターンに合致すれば提供済み
    - Drive:    設定の provided: true をそのまま提供済みとして扱う
                （Drive上の資料はローカルに無いため、台帳側で状態を持つ）
    """
    files = scan_inbox() if files is None else files
    results = []
    for req in data_requirements():
        patterns = req.get("match", [])
        matched = [
            f["path"]
            for f in files
            if any(fnmatch.fnmatch(f["path"].lower(), p.lower()) for p in patterns)
        ]
        declared = bool(req.get("provided", False))
        results.append(
            {
                "id": req.get("id"),
                "title": req.get("title"),
                "category": req.get("category"),
                "priority": req.get("priority", "medium"),
                "why": req.get("why", ""),
                "format": req.get("format", ""),
                "cadence": req.get("cadence", ""),
                "how_to_export": req.get("how_to_export", ""),
                "drive": req.get("drive", ""),
                "satisfied": bool(matched) or declared,
                "source": "drive" if declared and not matched else "inbox",
                "matched_files": matched or ([req["drive"]] if declared and req.get("drive") else []),
            }
        )
    return results


def documents_as_prompt(documents: list[dict]) -> str:
    """Claude に渡す資料ブロックを組み立てる。"""
    if not documents:
        return "（提供された資料はまだありません）"
    blocks = []
    for doc in documents:
        header = (
            f"### ファイル: {doc['path']}\n"
            f"- 分類: {doc['category']}（{doc['category_label']}）\n"
            f"- 更新日時: {doc['modified_at']}\n"
            f"- 読み取り: {doc['status']}"
            + ("（文字数上限で末尾を省略）" if doc["truncated"] else "")
        )
        body = doc["text"] if doc["status"] == "ok" else "（このファイルは本文を読み取れませんでした）"
        blocks.append(f"{header}\n\n```\n{body}\n```")
    return "\n\n".join(blocks)


def inbox_summary() -> dict:
    files = scan_inbox()
    reqs = requirement_status(files)
    return {
        "files": files,
        "file_count": len(files),
        "requirements": reqs,
        "satisfied_count": sum(1 for r in reqs if r["satisfied"]),
        "requirement_count": len(reqs),
        "categories": CATEGORIES,
    }
